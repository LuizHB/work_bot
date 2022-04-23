from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import time

print("Initiating the bot...\n")

domains = []

#file to save the results
file = open("Files/results.txt","w")

#sheet index
n = 0

#Reading from Excel file
PATH = "C:/Users/Pc/Documents/GitHub/work_bot/Bots/Files/Domains.xlsx"
workbook = load_workbook(PATH)
sheets = workbook.sheetnames
wb = workbook[sheets[n]]

#reading multiples rows and columns
for row in range(1,10):
    for column in range(1,3):
        print(wb.cell(row,column).value)
        domains.append(wb.cell(row,column).value)

#webdriver
ser = Service()
op = webdriver.ChromeOptions()
#option to remove bluetooth adapter error
op.add_experimental_option("excludeSwitches", ["enable-logging"])
driver = webdriver.Chrome(service=ser, options=op)
url = "https://registro.br/"
driver.get(url)

#domains list to check
for domain in domains:
    search = driver.find_element(By.ID, "is-avail-field")
    search.clear()  # Cleaning the search bar
    search.send_keys(domain)
    search.send_keys(Keys.RETURN)
    time.sleep(2) #time to load the page
    #find information with a Tag
    results = driver.find_elements(By.TAG_NAME,"strong")
    #saving results
    text ="Domain %s %s\n" % (domain, results[4].text)
    file.write(text)

file.close()
driver.close()
