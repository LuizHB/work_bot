from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import time

print("Initiating the bot...\n")

domains = []

#sheet index
n = 0

#Reading from Excel file
PATH = "C:/Users/Pc/Documents/GitHub/work_bot/Bots/ExcelFiles/Domains.xlsx"
workbook = load_workbook(PATH)
sheets = workbook.sheetnames
wb = workbook[sheets[n]]
for row in range(1,10):
    domains.append(wb.cell(row,1).value)

#webdriver
ser = Service('C:/Users/Pc/Desktop/Luiz/Robos/chromedriver')
op = webdriver.ChromeOptions()
#option to remove bluetooth adapter error
op.add_experimental_option("excludeSwitches", ["enable-logging"])
driver = webdriver.Chrome(service=ser, options=op)
url = "https://registro.br/"
driver.get(url)

#domains list to check
for domain in domains:
    search = driver.find_element(By.ID, "is-avail-field")
    search.clear()  # Cleaning the search bar
    search.send_keys(domain)
    search.send_keys(Keys.RETURN)
    time.sleep(2) #time to load the page
    #find information with a Tag
    results = driver.find_elements(By.TAG_NAME,"strong")
    #print results
    print("Domain %s %s" % (domain, results[4].text))

driver.close()

